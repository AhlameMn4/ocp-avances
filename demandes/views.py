from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse
from django.db.models import Count, Q
from accounts.decorators import admin_required, gestionnaire_required
from accounts.models import ActionLog
from events.models import Event
from agents.models import Agent
from .models import Demande


def _statut(event):
    return 'INCLUS' if event.type_avance == Event.TYPE_PREAMENAGEMENT else 'EXCLUS'


@gestionnaire_required
def saisie(request, event_pk):
    ev = get_object_or_404(Event, pk=event_pk)
    if not ev.est_actif:
        messages.info(request, "Cet événement est clôturé. La saisie n'est plus possible.")
        return redirect('events:gestionnaire_events')

    lbl, color, desc = ev.logique
    mes_demandes = Demande.objects.filter(
        event=ev, gestionnaire=request.user
    ).select_related('agent').order_by('-date_saisie')

    return render(request, 'demandes/saisie.html', {
        'ev': ev, 'mes_demandes': mes_demandes,
        'logique_label': lbl, 'logique_color': color, 'logique_desc': desc,
        'total': mes_demandes.count(),
    })


@gestionnaire_required
def confirmer_saisie(request, event_pk):
    if request.method != 'POST':
        return redirect('demandes:saisie', event_pk=event_pk)

    ev = get_object_or_404(Event, pk=event_pk)
    if not ev.est_actif:
        messages.info(request, "Cet événement est clôturé.")
        return redirect('events:gestionnaire_events')

    matricule = request.POST.get('matricule', '').strip().upper()
    action    = request.POST.get('action', '')

    try:
        agent = Agent.objects.get(matricule=matricule)
    except Agent.DoesNotExist:
        messages.error(request, f"Matricule {matricule} introuvable.")
        return redirect('demandes:saisie', event_pk=event_pk)

    if action == 'confirmer':
        if Demande.objects.filter(agent=agent, event=ev).exists():
            messages.warning(request, f"{agent.nom_complet} a déjà une demande pour cet événement.")
        else:
            statut = _statut(ev)
            Demande.objects.create(agent=agent, event=ev,
                                   gestionnaire=request.user, statut=statut)
            ActionLog.log(request, ActionLog.ACTION_SAISIE,
                          f"{agent.nom_complet} – {ev} [{statut}]")
            messages.success(request,
                f"✓ {agent.nom_complet} enregistré [{statut}]")
        return redirect('demandes:saisie', event_pk=event_pk)

    # Aperçu avant confirmation
    doublon = Demande.objects.filter(agent=agent, event=ev).exists()
    lbl, color, desc = ev.logique
    return render(request, 'demandes/confirmation.html', {
        'ev': ev, 'agent': agent, 'doublon': doublon,
        'statut_prevu': _statut(ev),
        'logique_label': lbl, 'logique_color': color,
    })


@admin_required
def liste_admin(request):
    qs = Demande.objects.select_related('agent', 'event', 'gestionnaire').all()
    from accounts.models import CustomUser

    event_id    = request.GET.get('event', '')
    service     = request.GET.get('service', '')
    gestionnaire= request.GET.get('gestionnaire', '')
    statut      = request.GET.get('statut', '')
    type_avance = request.GET.get('type_avance', '')
    q           = request.GET.get('q', '').strip()

    if event_id:     qs = qs.filter(event_id=event_id)
    if service:      qs = qs.filter(agent__service__icontains=service)
    if gestionnaire: qs = qs.filter(gestionnaire_id=gestionnaire)
    if statut:       qs = qs.filter(statut=statut)
    if type_avance:  qs = qs.filter(event__type_avance=type_avance)
    if q:
        qs = qs.filter(
            Q(agent__matricule__icontains=q) |
            Q(agent__nom__icontains=q) |
            Q(agent__prenom__icontains=q)
        )

    return render(request, 'demandes/liste_admin.html', {
        'demandes': qs,
        'events': Event.objects.all().order_by('-date_debut'),
        'gestionnaires': CustomUser.objects.filter(role=CustomUser.ROLE_GESTIONNAIRE),
        'services': Agent.objects.values_list('service', flat=True).distinct().order_by('service'),
        'total': qs.count(),
        'inclus': qs.filter(statut='INCLUS').count(),
        'exclus': qs.filter(statut='EXCLUS').count(),
        'filtres': request.GET,
    })

@admin_required
def liste_admin(request):
    from django.core.paginator import Paginator
    from accounts.models import CustomUser
    qs = Demande.objects.select_related('agent', 'event', 'gestionnaire').all()

    event_id     = request.GET.get('event', '')
    service      = request.GET.get('service', '')
    gestionnaire = request.GET.get('gestionnaire', '')
    statut       = request.GET.get('statut', '')
    type_avance  = request.GET.get('type_avance', '')
    q            = request.GET.get('q', '').strip()

    if event_id:     qs = qs.filter(event_id=event_id)
    if service:      qs = qs.filter(agent__service__icontains=service)
    if gestionnaire: qs = qs.filter(gestionnaire_id=gestionnaire)
    if statut:       qs = qs.filter(statut=statut)
    if type_avance:  qs = qs.filter(event__type_avance=type_avance)
    if q:
        qs = qs.filter(
            Q(agent__matricule__icontains=q) |
            Q(agent__nom__icontains=q) |
            Q(agent__prenom__icontains=q)
        )

    # Pagination 25 par page
    paginator = Paginator(qs, 25)
    page_num  = request.GET.get('page', 1)
    page_obj  = paginator.get_page(page_num)

    # Conserver les filtres dans les liens de pagination
    filtres_get = request.GET.copy()
    filtres_get.pop('page', None)

    return render(request, 'demandes/liste_admin.html', {
        'page_obj':      page_obj,
        'events':        Event.objects.all().order_by('-date_debut'),
        'gestionnaires': CustomUser.objects.filter(role=CustomUser.ROLE_GESTIONNAIRE),
        'services':      Agent.objects.values_list('service', flat=True).distinct().order_by('service'),
        'total':         qs.count(),
        'inclus':        qs.filter(statut='INCLUS').count(),
        'exclus':        qs.filter(statut='EXCLUS').count(),
        'filtres':       request.GET,
        'filtres_query': filtres_get.urlencode(),
    })


@admin_required
def export_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    qs = Demande.objects.select_related('agent', 'event', 'gestionnaire').all()
    if request.GET.get('event'):     qs = qs.filter(event_id=request.GET['event'])
    if request.GET.get('service'):   qs = qs.filter(agent__service__icontains=request.GET['service'])
    if request.GET.get('statut'):    qs = qs.filter(statut=request.GET['statut'])
    if request.GET.get('type_avance'): qs = qs.filter(event__type_avance=request.GET['type_avance'])

    ActionLog.log(request, ActionLog.ACTION_EXPORT, f"Export Excel – {qs.count()} demandes")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Avances OCP"

    ocp  = "005EB8"
    lite = "E8F0FA"
    hfont = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hfill = PatternFill("solid", fgColor=ocp)
    thin  = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A1:I1')
    ws['A1'].value = f"OCP Group – Avances Sociales – Export {timezone.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font  = Font(bold=True, color=ocp, name="Arial", size=12)
    ws['A1'].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28
    ws.append([])

    headers = ['Matricule','Nom','Prénom','Service','Type Avance','Année','Statut','Gestionnaire','Date Saisie']
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col)
        c.font=hfont; c.fill=hfill; c.alignment=center; c.border=thin
    ws.row_dimensions[3].height = 22

    for i, d in enumerate(qs, 1):
        row = [d.agent.matricule, d.agent.nom, d.agent.prenom, d.agent.service,
               d.event.get_type_avance_display(), d.event.annee,
               d.get_statut_display(), d.gestionnaire.get_full_name(),
               d.date_saisie.strftime('%d/%m/%Y %H:%M')]
        ws.append(row)
        ri = 3 + i
        fill = PatternFill("solid", fgColor=lite if i%2==0 else "FFFFFF")
        for col in range(1, 10):
            c = ws.cell(row=ri, column=col)
            c.font=Font(name="Arial",size=9); c.alignment=center; c.border=thin
            if col != 7: c.fill = fill
        sc = ws.cell(row=ri, column=7)
        sc.fill = PatternFill("solid", fgColor="D4EDDA" if d.statut=='INCLUS' else "F8D7DA")
        sc.font = Font(bold=True, name="Arial", size=9,
                       color="155724" if d.statut=='INCLUS' else "721c24")

    for i, w in enumerate([14,18,18,24,22,8,10,22,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="avances_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response